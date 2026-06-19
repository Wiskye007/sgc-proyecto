from flask import Blueprint, request, jsonify
from db_connection import db
import logging
from datetime import datetime

bp = Blueprint('reportes', __name__, url_prefix='/api/reportes')
logger = logging.getLogger(__name__)


@bp.route('/poblacion', methods=['GET'])
def get_poblacion():
    """Reporte de población por pabellón"""
    try:
        query = """
                SELECT Pabellon, COUNT(*) as cantidad
                FROM tblConvictos
                GROUP BY Pabellon
                ORDER BY Pabellon
                """
        result = db.execute_query(query)
        
        # Normalizar resultado
        poblacion_normalizada = []
        if result:
            for row in result:
                try:
                    pabellon = row.get('Pabellon') or row.get('pabellon') or 'Desconocido'
                    cantidad = int(row.get('cantidad') or 0)
                    poblacion_normalizada.append({
                        'pabellon': pabellon,
                        'cantidad': cantidad
                    })
                except (ValueError, TypeError) as e:
                    logger.warning(f"Error procesando fila de población: {row}, error: {e}")
                    continue
        
        return jsonify({'success': True, 'data': poblacion_normalizada}), 200
    except Exception as e:
        logger.error(f"Error en reporte de población: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'error': str(e), 'data': []}), 500


@bp.route('/incidentes-estadisticas', methods=['GET'])
def get_incidentes_estadisticas():
    """Estadísticas de incidentes (desde tblconducta)"""
    try:
        # Se actualiza el nombre de la tabla a tblconducta y la columna a 'tipo'
        query = """
                SELECT tipo, COUNT(*) as cantidad
                FROM tblconducta
                GROUP BY tipo
                ORDER BY cantidad DESC
                """
        result = db.execute_query(query)
        
        # Normalizar resultado para el frontend
        incidentes_normalizados = []
        if result:
            for row in result:
                try:
                    # db_connection a veces capitaliza las llaves (Tipo, Cantidad)
                    tipo = row.get('tipo') or row.get('Tipo') or 'Otro'
                    cantidad = int(row.get('cantidad') or row.get('Cantidad') or 0)
                    
                    incidentes_normalizados.append({
                        'tipo_incidente': tipo, # El frontend espera esta llave
                        'cantidad': cantidad
                    })
                except (ValueError, TypeError) as e:
                    logger.warning(f"Error procesando fila de incidentes: {row}, error: {e}")
                    continue
        
        return jsonify({'success': True, 'data': incidentes_normalizados}), 200
    except Exception as e:
        logger.error(f"Error en estadísticas de incidentes: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'error': str(e), 'data': []}), 500


@bp.route('/salud-estadisticas', methods=['GET'])
def get_salud_estadisticas():
    """Estadísticas de salud"""
    try:
        query = """
                SELECT Diagnostico, COUNT(*) as cantidad
                FROM tblRevisionesMedicas
                GROUP BY Diagnostico
                ORDER BY cantidad DESC
                LIMIT 10
                """
        result = db.execute_query(query)
        
        # Normalizar resultado
        salud_normalizada = []
        if result:
            for row in result:
                try:
                    diagnostico = row.get('Diagnostico') or row.get('diagnostico') or 'Desconocido'
                    cantidad = int(row.get('cantidad') or 0)
                    salud_normalizada.append({
                        'diagnostico': diagnostico,
                        'cantidad': cantidad
                    })
                except (ValueError, TypeError) as e:
                    logger.warning(f"Error procesando fila de salud: {row}, error: {e}")
                    continue
        
        return jsonify({'success': True, 'data': salud_normalizada}), 200
    except Exception as e:
        logger.error(f"Error en estadísticas de salud: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'error': str(e), 'data': []}), 500


@bp.route('/debug-columns', methods=['GET'])
def debug_columns():
    """Debug: Retorna información de las columnas disponibles"""
    try:
        tables_to_check = ['tblConvictos', 'incidentes', 'tblRevisionesMedicas']
        debug_info = {}
        
        for table in tables_to_check:
            try:
                query = f"SELECT * FROM {table} LIMIT 1"
                result = db.execute_query(query)
                if result and len(result) > 0:
                    debug_info[table] = {
                        'columns': list(result[0].keys()),
                        'sample': result[0]
                    }
                else:
                    debug_info[table] = {'columns': [], 'note': 'No data found'}
            except Exception as e:
                debug_info[table] = {'error': str(e)}
        
        return jsonify({'success': True, 'debug': debug_info}), 200
    except Exception as e:
        logger.error(f"Error en debug: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/tendencia-poblacion', methods=['GET'])
def get_tendencia_poblacion():
    """Tendencia de población en el tiempo"""
    try:
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')

        query = """
                SELECT CAST(fecha_ingreso AS DATE) as fecha, COUNT(*) as cantidad
                FROM tblConvictos
                WHERE fecha_ingreso BETWEEN ? AND ?
                GROUP BY CAST(fecha_ingreso AS DATE)
                ORDER BY fecha
                """

        result = db.execute_query(query, (fecha_inicio, fecha_fin))
        return jsonify({'success': True, 'data': result or []}), 200
    except Exception as e:
        logger.error(f"Error en tendencia de población: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'error': str(e), 'data': []}), 500


@bp.route('/guardar', methods=['POST'])
def guardar_reporte():
    """Guardar un nuevo reporte"""
    try:
        data = request.get_json()

        query = """
                INSERT INTO Reportes (tipo_reporte, asunto, fecha, personal_cargo, cargo_personal, observaciones, datos,
                                    formato)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?) \
                """

        import json
        datos_json = json.dumps(data.get('datos', {}))

        db.execute_update(query, (
            data.get('tipoReporte'),
            data.get('asunto'),
            data.get('fecha'),
            data.get('personalCargo'),
            data.get('cargoPersonal'),
            data.get('observaciones'),
            datos_json,
            data.get('formato', 'pdf')
        ))

        return jsonify({'success': True, 'message': 'Reporte guardado correctamente'}), 201
    except Exception as e:
        logger.error(f"Error guardando reporte: {e}")
        return jsonify({'error': 'Error al guardar reporte'}), 500


@bp.route('/historial', methods=['GET'])
def get_historial():
    """Obtener historial de reportes"""
    try:
        query = "SELECT id, tipo_reporte, asunto, fecha, formato, datos FROM Reportes ORDER BY fecha DESC"
        result = db.execute_query(query)

        reportes = []
        for row in result:
            import json
            reportes.append({
                'id': row['id'],
                'tipoReporte': row['tipo_reporte'],
                'asunto': row['asunto'],
                'fecha': str(row['fecha']),
                'formato': row['formato'],
                'datos': json.loads(row['datos']) if isinstance(row['datos'], str) else row['datos']
            })

        return jsonify({'success': True, 'data': reportes}), 200
    except Exception as e:
        logger.error(f"Error obteniendo historial: {e}")
        return jsonify({'error': 'Error al obtener historial'}), 500


@bp.route('/<report_id>', methods=['DELETE'])
def eliminar_reporte(report_id):
    """Eliminar un reporte"""
    try:
        query = "DELETE FROM Reportes WHERE id = ?"
        db.execute_update(query, (report_id,))
        return jsonify({'success': True, 'message': 'Reporte eliminado'}), 200
    except Exception as e:
        logger.error(f"Error eliminando reporte: {e}")
        return jsonify({'error': 'Error al eliminar reporte'}), 500


@bp.route('/<report_id>/exportar/<formato>', methods=['GET'])
def exportar_reporte(report_id, formato):
    """Exportar reporte en diferentes formatos"""
    try:
        query = "SELECT asunto, datos FROM Reportes WHERE id = ?"
        result = db.execute_query(query, (report_id,))

        if not result:
            return jsonify({'error': 'Reporte no encontrado'}), 404

        import json
        row = result[0]
        asunto = row['asunto']
        datos = row['datos']
        datos_dict = json.loads(datos) if isinstance(datos, str) else datos

        if formato == 'pdf':
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
            from io import BytesIO

            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            story = []
            styles = getSampleStyleSheet()

            story.append(Paragraph(f"<b>{asunto}</b>", styles['Heading1']))
            story.append(Spacer(1, 12))

            for key, value in datos_dict.items():
                story.append(Paragraph(f"<b>{key}:</b> {value}", styles['Normal']))

            doc.build(story)
            buffer.seek(0)

            from flask import send_file
            return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=f'{asunto}.pdf')

        elif formato == 'excel':
            import openpyxl
            from io import BytesIO

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte"

            row = 1
            for key, value in datos_dict.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = value
                row += 1

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            from flask import send_file
            return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            as_attachment=True, download_name=f'{asunto}.xlsx')

        elif formato == 'csv':
            import csv
            from io import StringIO

            buffer = StringIO()
            writer = csv.writer(buffer)

            for key, value in datos_dict.items():
                writer.writerow([key, value])

            from flask import send_file
            from io import BytesIO

            bytes_buffer = BytesIO(buffer.getvalue().encode())
            return send_file(bytes_buffer, mimetype='text/csv', as_attachment=True, download_name=f'{asunto}.csv')

        return jsonify({'error': 'Formato no soportado'}), 400
    except Exception as e:
        logger.error(f"Error exportando reporte: {e}")
        return jsonify({'error': 'Error al exportar reporte'}), 500 
